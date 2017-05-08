from openpyxl import load_workbook
class TestCase(object):
    list_testcase=[]
    def ListCase(self,filename,sheetnames,list_testcase):
        self.filename=filename
        self.sheetnames=sheetnames
        self.list_testcase=list_testcase
        wb = load_workbook(filename)
        ws = wb.get_sheet_by_name(sheetnames)
        dict_testcase={}
        for nrow in range(2, ws.max_row+1 ):
            dict_testcase={}
            for ncol in range( 1, ws.max_column+1 ):
                key = ws.cell( row=1, column=ncol ).value
                value = ws.cell( row=nrow, column=ncol ).value
                dict_testcase[key]=value
            list_testcase.append(dict_testcase)
        list_testcase=TestCase.list_testcase
        return list_testcase
    

